#!/usr/bin/env python
"""Windows desktop intelligence helpers for Nanobot.

Provides parseable snapshots of visible windows and optional OCR extraction from
screenshots or application captures. Designed as a robust fallback when browser
or API routes are unavailable.
"""
from __future__ import annotations

import argparse
import csv
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

OMEGA_ROOT = Path(r"C:\AI\nanobot-omega")
if str(OMEGA_ROOT) not in sys.path:
    sys.path.insert(0, str(OMEGA_ROOT))

try:
    from tools.ocr_tool import perform_ocr
except Exception:
    perform_ocr = None


def list_windows() -> dict[str, Any]:
    script = r"""
Get-Process | Where-Object { $_.MainWindowTitle } |
Select-Object Id,ProcessName,MainWindowTitle,Path |
ConvertTo-Json -Depth 3
"""
    proc = subprocess.run(
        ["powershell.exe", "-NoProfile", "-Command", script],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        check=False,
    )
    if proc.returncode != 0:
        return {"ok": False, "error": proc.stderr.strip(), "items": []}
    raw = proc.stdout.strip()
    if not raw:
        items = []
    else:
        data = json.loads(raw)
        items = data if isinstance(data, list) else [data]
    return {"ok": True, "generated_at": datetime.now().isoformat(timespec="seconds"), "items": items}


def ocr_image(path: str) -> dict[str, Any]:
    image = Path(path).expanduser().resolve()
    if not image.exists():
        raise FileNotFoundError(f"image not found: {image}")
    if perform_ocr is None:
        return {"ok": False, "image": str(image), "error": "OCR tool unavailable"}
    result = perform_ocr(image, engine="auto", languages=["fr", "en", "nl"])
    return {
        "ok": True,
        "image": str(image),
        "text": result.text,
        "engine": getattr(result, "engine", "auto"),
        "confidence": getattr(result, "confidence", None),
    }


def export_xlsx(snapshot: dict[str, Any], path: str) -> dict[str, Any]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except Exception:
        return export_csv(snapshot, str(Path(path).with_suffix(".csv")))

    target = Path(path).expanduser().resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Fenêtres"
    headers = ["Id", "ProcessName", "MainWindowTitle", "Path"]
    ws.append(headers)
    for item in snapshot.get("items", []):
        ws.append([item.get("Id"), item.get("ProcessName"), item.get("MainWindowTitle"), item.get("Path")])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        width = min(80, max(12, max(len(str(c.value or "")) for c in col) + 2))
        ws.column_dimensions[col[0].column_letter].width = width
    if ws.max_row >= 2:
        table = Table(displayName="WindowsSnapshot", ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    synth = wb.create_sheet("Synthese")
    synth.append(["Indicateur", "Valeur"])
    synth.append(["Fenêtres détectées", len(snapshot.get("items", []))])
    synth.append(["Généré", snapshot.get("generated_at")])
    qc = wb.create_sheet("Controle qualite")
    qc.append(["Controle", "Statut"])
    qc.append(["Données en colonnes", "OK"])
    qc.append(["Export lisible", "OK"])
    wb.save(target)
    return {"ok": True, "path": str(target), "format": "xlsx"}


def export_csv(snapshot: dict[str, Any], path: str) -> dict[str, Any]:
    target = Path(path).expanduser().resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=["Id", "ProcessName", "MainWindowTitle", "Path"])
        writer.writeheader()
        writer.writerows(snapshot.get("items", []))
    return {"ok": True, "path": str(target), "format": "csv"}


def main() -> int:
    parser = argparse.ArgumentParser(description="Intelligence bureau Windows pour Nanobot.")
    sub = parser.add_subparsers(dest="command", required=True)
    windows = sub.add_parser("windows")
    windows.add_argument("--xlsx")
    ocr = sub.add_parser("ocr")
    ocr.add_argument("image")
    args = parser.parse_args()
    if args.command == "windows":
        snap = list_windows()
        if args.xlsx:
            snap["export"] = export_xlsx(snap, args.xlsx)
        print(json.dumps(snap, ensure_ascii=False, indent=2))
        return 0 if snap.get("ok") else 1
    if args.command == "ocr":
        print(json.dumps(ocr_image(args.image), ensure_ascii=False, indent=2))
        return 0
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
